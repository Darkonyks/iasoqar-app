"""
Django management command za import podataka kompanija i nadzornih provera iz Excel fajlova
"""
from django.core.management.base import BaseCommand
from django.core.management import call_command
from django.db import transaction
from company.models import (
    Company, CertificationCycle, CycleAudit, CycleStandard,
    StandardDefinition, IAFEACCode, CompanyIAFEACCode, Certificate
)
from datetime import datetime
import openpyxl
import os


class Command(BaseCommand):
    help = 'Importuje podatke kompanija i nadzornih provera iz Excel fajlova'

    def add_arguments(self, parser):
        parser.add_argument(
            'company_file',
            type=str,
            help='Putanja do Excel fajla sa kompanijama (company-list.xlsx)'
        )
        parser.add_argument(
            'provere_file',
            type=str,
            help='Putanja do Excel fajla sa nadzornim proverama (naredne-provere.xlsx)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Testiranje bez čuvanja u bazu'
        )
        parser.add_argument(
            '--limit',
            type=int,
            default=None,
            help='Broj redova za import (za testiranje)'
        )

    def handle(self, *args, **options):
        company_file = options['company_file']
        provere_file = options['provere_file']
        dry_run = options['dry_run']
        limit = options['limit']

        # Otvori log fajl
        log_file = open('import_log.txt', 'w', encoding='utf-8')
        
        def log(message):
            """Helper funkcija za logovanje u fajl i stdout"""
            self.stdout.write(message)
            log_file.write(message + '\n')
            log_file.flush()

        if not os.path.exists(company_file):
            log(self.style.ERROR(f'Fajl ne postoji: {company_file}'))
            log_file.close()
            return

        if not os.path.exists(provere_file):
            log(self.style.ERROR(f'Fajl ne postoji: {provere_file}'))
            log_file.close()
            return

        if dry_run:
            log(self.style.WARNING(' DRY RUN MODE - Nema promene u bazi'))

        # Sačuvaj log funkciju kao atribut
        self.log = log

        # Učitaj standarde ako ne postoje
        if StandardDefinition.objects.count() == 0:
            log(self.style.WARNING('⚠️  Nema standarda u bazi, učitavam fixture...'))
            try:
                call_command('loaddata', 'standard_definitions', verbosity=0)
                log(self.style.SUCCESS(f'✅ Učitano {StandardDefinition.objects.count()} standarda'))
            except Exception as e:
                log(self.style.ERROR(f'Greška pri učitavanju standarda: {str(e)}'))

        try:
            with transaction.atomic():
                # Prvo importuj kompanije
                companies_map = self.import_companies(company_file, dry_run, limit)
                
                # Zatim importuj nadzorne provere
                self.import_nadzorne_provere(provere_file, companies_map, dry_run)
                
                if dry_run:
                    raise Exception("Dry run - rollback transakcije")
                    
        except Exception as e:
            if not dry_run:
                log(self.style.ERROR(f'Greška: {str(e)}'))
            else:
                log(self.style.SUCCESS('Dry run završen uspešno!'))
        finally:
            log('\n✅ Import završen! Log sačuvan u import_log.txt')
            log_file.close()

    def parse_date(self, date_value):
        """Parsira datum iz različitih formata"""
        if not date_value:
            return None
        
        if isinstance(date_value, datetime):
            return date_value.date()
        
        if isinstance(date_value, str):
            # Pokušaj različite formate
            for fmt in ['%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                try:
                    return datetime.strptime(date_value, fmt).date()
                except ValueError:
                    continue
        
        return None

    def import_companies(self, file_path, dry_run, limit=None):
        """
        Importuje kompanije iz Excel fajla (iz svih sheet-ova)
        Vraća dict: {company_id: Company instance}
        """
        self.stdout.write(self.style.SUCCESS(f'\n📊 Učitavanje kompanija iz: {file_path}'))
        if limit:
            self.stdout.write(f'   Limit: prvih {limit} redova po sheet-u')
        
        wb = openpyxl.load_workbook(file_path)
        
        companies_map = {}
        created_count = 0
        updated_count = 0
        skipped_count = 0
        
        # Učitaj iz Sheet1 i dupli
        sheet_names = ['Sheet1', 'dupli']
        self.stdout.write(f'Importovanje iz sheet-ova: {sheet_names}')
        
        # Učitaj kompanije iz svih sheet-ova
        for sheet_name in sheet_names:
            self.stdout.write(f'\n📄 Učitavanje kompanija iz sheet-a: {sheet_name}')
            
            ws = wb[sheet_name]
            
            # Pretpostavljamo da je prvi red header
            headers = [cell.value for cell in ws[1]]
            self.stdout.write(f'  Kolone: {headers}')
            
            sheet_created = 0
            sheet_updated = 0
            sheet_skipped = 0
            
            rows_processed = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Proveri limit
                if limit and rows_processed >= limit:
                    self.stdout.write(f'    Dostignut limit od {limit} redova')
                    break
                rows_processed += 1
                
                try:
                    # Mapiranje kolona prema stvarnoj strukturi:
                    # ['company_id', 'company_name', 'certificate_no', 'init_reg_date', 
                    #  'standard', 'certificate_status', 'suspension_until_date', 'audit_days', 
                    #  'initial_audit_conducted_date', 'visits_per_year', 'audit_days_each']
                    company_id = row[0]  # company_id
                    company_name = row[1]  # company_name
                    certificate_number = row[2] if len(row) > 2 else None  # certificate_no
                    init_reg_date = self.parse_date(row[3]) if len(row) > 3 else None  # init_reg_date
                    standard_codes = row[4] if len(row) > 4 else None  # standard (može biti '9,14,18')
                    certificate_status = row[5] if len(row) > 5 else 'active'  # certificate_status
                    suspension_until_date = self.parse_date(row[6]) if len(row) > 6 else None
                    audit_days = row[7] if len(row) > 7 else None
                    initial_audit_conducted_date = self.parse_date(row[8]) if len(row) > 8 else None
                    visits_per_year = row[9] if len(row) > 9 else None
                    audit_days_each = row[10] if len(row) > 10 else None
                    
                    if not company_name:
                        sheet_skipped += 1
                        continue
                
                    # Mapiranje statusa
                    status_map = {
                        'ACTIVE': 'active',
                        'SUSPENDED': 'suspended',
                        'WITHDRAWN': 'withdrawn',
                        'EXPIRED': 'expired',
                        'PENDING': 'pending',
                        'CANCELLED': 'cancelled',
                    }
                    cert_status = status_map.get(str(certificate_status).upper(), 'active')
                    
                    # Kreiraj ili ažuriraj kompaniju
                    # VAŽNO: Prvo proveravamo da li postoji sertifikat sa istim brojem
                    # jer ista kompanija može imati različite nazive u Excel-u
                    # (npr. "ADAM ŠPED SYSTEM" i "ADAM-ŠPED SYSTEM DOO" su ista kompanija)
                    company = None
                    certificate = None
                    created = False
                    
                    if certificate_number:
                        # Prvo tražimo sertifikat po broju
                        certificate = Certificate.objects.filter(
                            certificate_number=str(certificate_number)
                        ).first()
                        if certificate:
                            company = certificate.company
                    
                    if not company:
                        # Ako nema po broju sertifikata, tražimo kompaniju po imenu
                        company = Company.objects.filter(name=company_name).first()
                    
                    if company:
                        # Kompanija već postoji
                        # Ažuriraj ili kreiraj sertifikat za ovu kompaniju
                        if certificate_number and not certificate:
                            # Kreiraj novi sertifikat za postojeću kompaniju
                            certificate, cert_created = Certificate.objects.get_or_create(
                                certificate_number=str(certificate_number),
                                defaults={
                                    'company': company,
                                    'status': cert_status,
                                    'suspension_until_date': suspension_until_date,
                                }
                            )
                            if not cert_created and sheet_name.lower() not in ['dupli', 'duplicate', 'duplicates']:
                                # Ažuriraj postojeći sertifikat
                                certificate.status = cert_status
                                certificate.suspension_until_date = suspension_until_date or certificate.suspension_until_date
                                certificate.save()
                        elif certificate and sheet_name.lower() not in ['dupli', 'duplicate', 'duplicates']:
                            # Ažuriraj postojeći sertifikat
                            certificate.status = cert_status
                            certificate.suspension_until_date = suspension_until_date or certificate.suspension_until_date
                            certificate.save()
                        sheet_updated += 1
                    else:
                        # Kreiraj novu kompaniju
                        company = Company.objects.create(
                            name=company_name,
                        )
                        created = True
                        sheet_created += 1
                        
                        # Kreiraj sertifikat za novu kompaniju
                        if certificate_number:
                            certificate = Certificate.objects.create(
                                company=company,
                                certificate_number=str(certificate_number),
                                status=cert_status,
                                suspension_until_date=suspension_until_date,
                            )
                    
                    # Čuvamo kompaniju i dodatne podatke za ciklus
                    # VAŽNO: standard_codes se čuva da bi se znalo koji standardi pripadaju ovom company_id
                    companies_map[company_id] = {
                        'company': company,
                        'audit_days': audit_days,  # inicijalni_broj_dana
                        'visits_per_year': visits_per_year,  # broj_dana_nadzora
                        'audit_days_each': audit_days_each,  # broj_dana_resertifikacije
                        'initial_audit_conducted_date': initial_audit_conducted_date,
                        'standard_codes': standard_codes,  # standardi specifični za ovaj company_id
                    }
                    
                    # Dodaj standarde ako su navedeni (mogu biti odvojeni zarezom: '9,14,18')
                    if standard_codes:
                        self.add_standards_to_company(company, standard_codes, init_reg_date)
                    
                    # Za sheet "dupli" - NE kreiramo cikluse ovde jer se oni kreiraju iz naredne-provere.xlsx
                    # Sheet "dupli" samo dodaje standarde kompaniji i čuva company_id mapiranje
                    
                    if row_idx % 10 == 0:
                        self.stdout.write(f'    Obrađeno {row_idx} redova...')
                        
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f'    Greška u redu {row_idx}: {str(e)}'))
                    continue
            
            # Prikaz statistike za sheet
            created_count += sheet_created
            updated_count += sheet_updated
            skipped_count += sheet_skipped
            self.stdout.write(self.style.SUCCESS(
                f'  ✅ Sheet "{sheet_name}": {sheet_created} kreirano, {sheet_updated} ažurirano, {sheet_skipped} preskočeno'
            ))
        
        self.stdout.write(self.style.SUCCESS(
            f'\n✅ Kompanije: {created_count} kreirano, {updated_count} ažurirano, {skipped_count} preskočeno'
        ))
        
        return companies_map

    def add_standards_to_company(self, company, standard_codes, issue_date=None):
        """
        Dodaje standarde kompaniji
        standard_codes može biti: '9,14,18' ili '9.14' ili '9001;27001' ili '9' ili broj (float 9.14)
        ili veliki broj bez separatora: 90011400118001 (Excel tretira kao int)
        issue_date: Datum izdavanja sertifikata (init_reg_date iz Excel-a)
        """
        try:
            # Konvertuj u string (ako je broj kao 9.14, postaje "9.14")
            standard_codes = str(standard_codes) if standard_codes else ''
            
            # Proveri da li je ovo veliki broj bez separatora (npr. 90011400118001)
            # Ako je string samo cifre i duži od 5 karaktera, razbij ga na delove
            if standard_codes.isdigit() and len(standard_codes) > 5:
                # Razbij na delove od 4-5 cifara (ISO standardi: 9001, 14001, 18001, 45001, 22000, 27001, 20000, 50001)
                codes = []
                i = 0
                while i < len(standard_codes):
                    # Pokušaj da uzmeš 5 cifara (npr. 45001, 22000, 22301, 3834)
                    if i + 5 <= len(standard_codes):
                        code = standard_codes[i:i+5]
                        # Proveri da li je validan ISO kod sa 5 cifara
                        if code in ['45001', '22000', '27001', '20000', '50001', '22301']:
                            codes.append(code)
                            i += 5
                            continue
                    # Proveri i za 4-cifrene standarde koji mogu biti na početku (npr. 3834)
                    if i + 4 <= len(standard_codes):
                        code = standard_codes[i:i+4]
                        # Ako je 3834 ili drugi specifičan 4-cifreni kod
                        if code == '3834':
                            codes.append(code)
                            i += 4
                            continue
                    # Inače uzmi 4 cifre (npr. 9001, 14001, 18001)
                    if i + 4 <= len(standard_codes):
                        code = standard_codes[i:i+4]
                        codes.append(code)
                        i += 4
                    else:
                        # Preostale cifre (ne bi trebalo da se desi)
                        codes.append(standard_codes[i:])
                        break
            else:
                # Normalizuj separatore (tačke i tačka-zarezi → zarezi)
                standard_codes = standard_codes.replace('.', ',').replace(';', ',')
                
                # Ako ima zareze, razdvoji
                if ',' in standard_codes:
                    codes = [code.strip() for code in standard_codes.split(',')]
                # Ako ima razmake (npr. "9 14 45 22301"), razdvoji po razmaku
                elif ' ' in standard_codes:
                    codes = [code.strip() for code in standard_codes.split(' ') if code.strip()]
                else:
                    codes = [standard_codes.strip()]
            
            for code in codes:
                if code:  # Preskoči prazne stringove
                    self.add_standard_to_company(company, code, issue_date)
                
        except Exception as e:
            self.stdout.write(self.style.WARNING(f'    Greška pri dodavanju standarda: {str(e)}'))
    
    def add_standard_to_company(self, company, standard_id, issue_date=None):
        """
        Dodaje jedan standard kompaniji
        Mapira skraćene kodove: 9 → 9001, 14 → 14001, 18 → 18001, 45 → 45001, itd.
        issue_date: Datum izdavanja sertifikata
        """
        try:
            # Pokušaj da pronađeš standard po ID-u ili kodu
            standard_def = None
            
            # Mapiranje skraćenih kodova na kodove u bazi (sa ISO prefiksom)
            standard_mapping = {
                '9': 'ISO9001',
                '9001': 'ISO9001',
                '14': 'ISO14001',
                '14001': 'ISO14001',
                '18': 'ISO45001',  # Stari OHSAS 18001 → mapira na ISO 45001
                '18001': 'ISO45001',
                '45': 'ISO45001',
                '45001': 'ISO45001',
                '22': 'ISO22000',
                '22000': 'ISO22000',
                '27': 'ISO27001',
                '27001': 'ISO27001',
                '20': 'ISO20000',
                '20000': 'ISO20000',
                '50': 'ISO50001',
                '50001': 'ISO50001',
                '22301': 'ISO22301',
                '37001': 'ISO37001',
                '13485': 'ISO13485',
                'HACCP': 'HACCP',
            }
            
            # Konvertuj u string
            std_code = str(standard_id).strip()
            original_code = std_code
            
            # Ako je skraćeni kod, mapiraj ga na puni
            if std_code in standard_mapping:
                std_code = standard_mapping[std_code]
            
            # Ako je standard_id broj, pokušaj da nađeš po ID-u
            if std_code.isdigit() and len(std_code) < 4:
                try:
                    standard_def = StandardDefinition.objects.get(id=int(std_code))
                except StandardDefinition.DoesNotExist:
                    pass
            
            # Ako nije pronađen, pokušaj po kodu
            if not standard_def:
                try:
                    # Prvo pokušaj tačno poklapanje
                    standard_def = StandardDefinition.objects.filter(
                        code__iexact=std_code
                    ).first()
                    
                    # Pokušaj sa 'ISO ' prefiksom (startswith)
                    if not standard_def:
                        standard_def = StandardDefinition.objects.filter(
                            code__istartswith=f'ISO {std_code}'
                        ).first()
                    
                    # Pokušaj bez razmaka (startswith)
                    if not standard_def:
                        standard_def = StandardDefinition.objects.filter(
                            code__istartswith=f'ISO{std_code}'
                        ).first()
                    
                    # Pokušaj contains kao poslednju opciju
                    if not standard_def:
                        standard_def = StandardDefinition.objects.filter(
                            code__icontains=std_code
                        ).first()
                        
                except StandardDefinition.DoesNotExist:
                    pass
            
            if standard_def:
                # Kreiraj vezu između kompanije i standarda
                from company.models import CompanyStandard
                company_standard, created = CompanyStandard.objects.get_or_create(
                    company=company,
                    standard_definition=standard_def,
                    defaults={'issue_date': issue_date}
                )
                if created:
                    self.stdout.write(f'      ✓ Dodat standard {standard_def.code} za {company.name}')
                # Ako već postoji, ažuriraj issue_date ako je prosleđen
                if not created and issue_date and not company_standard.issue_date:
                    company_standard.issue_date = issue_date
                    company_standard.save(update_fields=['issue_date'])
            else:
                self.stdout.write(self.style.WARNING(f'      ⚠ Standard "{original_code}" → "{std_code}" nije pronađen'))
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'      ✗ Greška pri dodavanju standarda: {str(e)}'))

    def add_iaf_code_to_company(self, company, iaf_kod_id):
        """Dodaje IAF/EAC kod kompaniji"""
        try:
            # Pokušaj da pronađeš IAF kod po ID-u ili kodu
            iaf_code = None
            
            if isinstance(iaf_kod_id, int) or (isinstance(iaf_kod_id, str) and iaf_kod_id.isdigit()):
                try:
                    iaf_code = IAFEACCode.objects.get(id=int(iaf_kod_id))
                except IAFEACCode.DoesNotExist:
                    pass
            
            if not iaf_code and iaf_kod_id:
                try:
                    iaf_code = IAFEACCode.objects.get(code__icontains=str(iaf_kod_id))
                except IAFEACCode.DoesNotExist:
                    self.stdout.write(self.style.WARNING(f'    IAF kod {iaf_kod_id} nije pronađen'))
                    return
            
            if iaf_code:
                CompanyIAFEACCode.objects.get_or_create(
                    company=company,
                    iaf_eac_code=iaf_code
                )
                
        except Exception as e:
            self.stdout.write(self.style.WARNING(f'    Greška pri dodavanju IAF koda: {str(e)}'))

    def create_certification_cycle(self, company, start_date, audit_days, visits_per_year, audit_days_each, standard_codes=None):
        """
        Kreira certification cycle za kompaniju
        Ako ima više standarda (npr. '9,14,18'), kreira integrisani sistem
        """
        try:
            # Proveri da li ima više standarda (integrisani sistem)
            is_integrated = False
            standards_list = []
            
            if standard_codes:
                # Konvertuj u string (ako je broj kao 9.14, postaje "9.14")
                standard_codes = str(standard_codes)
                
                # Normalizuj separatore (tačke i tačka-zarezi → zarezi)
                standard_codes = standard_codes.replace('.', ',').replace(';', ',')
                
                if ',' in standard_codes:
                    standards_list = [code.strip() for code in standard_codes.split(',') if code.strip()]
                    is_integrated = len(standards_list) > 1
                else:
                    standards_list = [standard_codes.strip()]
            
            # Proveri da li već postoji cycle za ovu kompaniju sa ovim datumom
            cycle, created = CertificationCycle.objects.get_or_create(
                company=company,
                planirani_datum=start_date,
                defaults={
                    'status': 'active',
                    'inicijalni_broj_dana': audit_days,
                    'datum_sprovodjenja_inicijalne': start_date,
                    'is_integrated_system': is_integrated,
                }
            )
            
            if created:
                # Dodaj standarde u cycle
                if standards_list:
                    self.add_standards_to_cycle(cycle, standards_list)
                
                # NE kreiramo audite ovde - auditi se kreiraju iz naredne-provere.xlsx
                
        except Exception as e:
            self.stdout.write(self.style.WARNING(f'    Greška pri kreiranju ciklusa: {str(e)}'))
    
    def create_cycle_from_company_row(self, company, init_reg_date, standard_codes, 
                                       audit_days, visits_per_year, audit_days_each,
                                       initial_audit_conducted_date, cert_status):
        """
        Kreira CertificationCycle direktno iz reda u company-list.xlsx (sheet dupli)
        Ovo se koristi za dodatne sertifikate iste kompanije
        """
        from company.models import CertificationCycle, CycleStandard, CompanyStandard
        
        try:
            # Parsiraj standarde
            standards_list = self.parse_standard_codes(standard_codes) if standard_codes else []
            
            # Kreiraj ciklus - koristimo init_reg_date kao planirani_datum
            # i initial_audit_conducted_date kao datum_sprovodjenja_inicijalne
            cycle, created = CertificationCycle.objects.get_or_create(
                company=company,
                planirani_datum=init_reg_date,
                defaults={
                    'status': 'active' if cert_status == 'active' else 'archived',
                    'inicijalni_broj_dana': audit_days,
                    'broj_dana_nadzora': visits_per_year,
                    'broj_dana_resertifikacije': audit_days_each,
                    'datum_sprovodjenja_inicijalne': initial_audit_conducted_date,
                }
            )
            
            if created:
                self.stdout.write(self.style.SUCCESS(
                    f'    ✅ Kreiran ciklus (dupli) za {company.name}, datum: {init_reg_date}, standardi: {standards_list}'
                ))
                
                # Dodaj standarde u ciklus
                if standards_list:
                    self.add_standards_to_cycle(cycle, standards_list)
                
                # Detektuj da li je integrisani sistem
                if len(standards_list) > 1:
                    cycle.is_integrated_system = True
                    cycle.save(update_fields=['is_integrated_system'])
            else:
                self.stdout.write(f'    ⚠️  Ciklus već postoji za {company.name}, datum: {init_reg_date}')
                
        except Exception as e:
            self.stdout.write(self.style.WARNING(f'    Greška pri kreiranju ciklusa (dupli): {str(e)}'))
    
    def parse_standard_codes(self, standard_codes):
        """Parsira string standarda u listu kodova"""
        if not standard_codes:
            return []
        
        standard_codes = str(standard_codes)
        
        # Razdvoji po zarezima, tačkama, tačka-zarezima ili razmacima
        import re
        codes = re.split(r'[,;.\s]+', standard_codes)
        return [c.strip() for c in codes if c.strip()]
    
    def add_standards_to_cycle(self, cycle, standards_list):
        """Dodaje standarde u certification cycle"""
        from company.models import CycleStandard
        
        for std_code in standards_list:
            # Mapiranje skraćenih kodova na kodove u bazi (sa ISO prefiksom)
            standard_mapping = {
                '9': 'ISO9001',
                '9001': 'ISO9001',
                '14': 'ISO14001',
                '14001': 'ISO14001',
                '18': 'ISO45001',
                '18001': 'ISO45001',
                '45': 'ISO45001',
                '45001': 'ISO45001',
                '22': 'ISO22000',
                '22000': 'ISO22000',
                '27': 'ISO27001',
                '27001': 'ISO27001',
                '20': 'ISO20000',
                '20000': 'ISO20000',
                '50': 'ISO50001',
                '50001': 'ISO50001',
                '22301': 'ISO22301',
                '37001': 'ISO37001',
                '13485': 'ISO13485',
                'HACCP': 'HACCP',
            }
            
            std_code = str(std_code).strip()
            if std_code in standard_mapping:
                std_code = standard_mapping[std_code]
            
            # Pronađi StandardDefinition
            try:
                # Prvo pokušaj tačno poklapanje
                standard_def = StandardDefinition.objects.filter(
                    code__iexact=std_code
                ).first()
                
                # Pokušaj sa 'ISO ' prefiksom
                if not standard_def:
                    standard_def = StandardDefinition.objects.filter(
                        code__istartswith=f'ISO {std_code}'
                    ).first()
                
                # Pokušaj bez razmaka
                if not standard_def:
                    standard_def = StandardDefinition.objects.filter(
                        code__istartswith=f'ISO{std_code}'
                    ).first()
                
                # Pokušaj contains (npr. '9001' u 'ISO 9001:2015')
                if not standard_def:
                    standard_def = StandardDefinition.objects.filter(
                        code__icontains=std_code
                    ).first()
                
                if standard_def:
                    # Dodaj standard u cycle
                    CycleStandard.objects.get_or_create(
                        certification_cycle=cycle,
                        standard_definition=standard_def  # Ispravljeno: koristi standard_definition
                    )
            except Exception as e:
                pass  # Tiho ignoriši greške

    def import_nadzorne_provere(self, file_path, companies_map, dry_run):
        """
        Importuje nadzorne provere iz Excel fajla
        """
        self.stdout.write(self.style.SUCCESS(f'\n📊 Učitavanje nadzornih provera iz: {file_path}'))
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        created_count = 0
        skipped_count = 0
        
        # Pretpostavljamo da je prvi red header
        headers = [cell.value for cell in ws[1]]
        self.stdout.write(f'Kolone: {headers}')
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                naredne_provere_id = row[0]
                company_id = row[1]
                first_surv_due = self.parse_date(row[2]) if len(row) > 2 else None
                first_surv_cond = self.parse_date(row[3]) if len(row) > 3 else None
                second_surv_due = self.parse_date(row[4]) if len(row) > 4 else None
                second_surv_cond = self.parse_date(row[5]) if len(row) > 5 else None
                trinial_audit_due = self.parse_date(row[6]) if len(row) > 6 else None
                trinial_audit_cond = self.parse_date(row[7]) if len(row) > 7 else None
                status_id = row[8] if len(row) > 8 else None
                
                # Pronađi kompaniju i njene podatke
                company_data = companies_map.get(company_id)
                if not company_data:
                    self.stdout.write(self.style.WARNING(
                        f'  Red {row_idx}: Kompanija sa ID {company_id} nije pronađena'
                    ))
                    skipped_count += 1
                    continue
                
                company = company_data['company']
                audit_days = company_data.get('audit_days')
                visits_per_year = company_data.get('visits_per_year')
                audit_days_each = company_data.get('audit_days_each')
                
                # Svaki red u naredne-provere.xlsx predstavlja JEDAN ciklus sertifikacije
                # Kreiraj novi cycle za ovaj red (ne tražimo postojeći)
                # Koristimo prvi nadzor kao planirani datum ciklusa
                cycle_date = first_surv_due or second_surv_due or trinial_audit_due
                if not cycle_date:
                    self.stdout.write(self.style.WARNING(
                        f'  Red {row_idx}: Nema datuma za kreiranje ciklusa'
                    ))
                    skipped_count += 1
                    continue
                
                # Mapiranje statusa
                # VAŽNO: Ciklus je archived SAMO ako su SVA TRI datuma završetka uneta
                # i nisu placeholder vrednosti (0001-01-01)
                # Placeholder datum 0001-01-01 se koristi za NULL vrednosti u Excel-u
                status_mapping = {
                    'ACTIVE': 'active',
                    'ARCHIVE': 'archived',
                }
                
                # Helper funkcija za proveru da li je datum validan (nije NULL i nije placeholder)
                def is_valid_date(d):
                    if not d:
                        return False
                    # Proveri da li je placeholder datum (0001-01-01)
                    if hasattr(d, 'year') and d.year == 1:
                        return False
                    return True
                
                # Ciklus je završen SAMO ako su sva tri datuma završetka validna
                all_audits_completed = (
                    is_valid_date(first_surv_cond) and 
                    is_valid_date(second_surv_cond) and 
                    is_valid_date(trinial_audit_cond)
                )
                
                if all_audits_completed:
                    # Svi auditi su sprovedeni - ciklus je završen
                    cycle_status = 'archived'
                else:
                    # Koristi status iz Excel-a
                    cycle_status = status_mapping.get(str(status_id).upper(), 'active')
                
                # Kreiraj ili pronađi cycle za ovaj red
                # Koristimo get_or_create da izbegnemo duplikate
                cycle, created = CertificationCycle.objects.get_or_create(
                    company=company,
                    planirani_datum=cycle_date,
                    defaults={
                        'status': cycle_status,
                        'inicijalni_broj_dana': audit_days,
                        'broj_dana_nadzora': visits_per_year,
                        'broj_dana_resertifikacije': audit_days_each,
                    }
                )
                
                if created:
                    self.log(f'  ✅ Kreiran ciklus za {company.name}, datum: {cycle_date}, status: {cycle_status}')
                else:
                    self.log(f'  ⚠️  Ciklus već postoji za {company.name}, datum: {cycle_date}')
                
                # Ako već postoji, ažuriraj status
                if not created and cycle.status != cycle_status:
                    cycle.status = cycle_status
                    cycle.save(update_fields=['status'])
                
                # Dodaj standarde u ciklus - SAMO standarde specifične za ovaj company_id
                # (ne sve standarde kompanije, jer kompanija može imati više sertifikata)
                if created:
                    from company.models import CycleStandard
                    standard_codes = company_data.get('standard_codes')
                    if standard_codes:
                        # Parsiraj i dodaj standarde specifične za ovaj company_id
                        standards_list = self.parse_standard_codes(standard_codes)
                        self.add_standards_to_cycle(cycle, standards_list)
                        
                        # Detektuj da li je integrisani sistem
                        if len(standards_list) > 1:
                            cycle.is_integrated_system = True
                            cycle.save(update_fields=['is_integrated_system'])
                
                # Kreiraj prvi nadzorni audit
                # VAŽNO: Onemogućavamo automatsko kreiranje drugog nadzora tokom importa
                if first_surv_due:
                    existing_audit = CycleAudit.objects.filter(
                        certification_cycle=cycle,
                        audit_type='surveillance_1',
                        planned_date=first_surv_due
                    ).first()
                    
                    if existing_audit:
                        if first_surv_cond and existing_audit.actual_date != first_surv_cond:
                            existing_audit._skip_cycle_creation = True
                            existing_audit.actual_date = first_surv_cond
                            existing_audit.audit_status = 'completed'
                            existing_audit.save()
                    else:
                        audit = CycleAudit(
                            certification_cycle=cycle,
                            audit_type='surveillance_1',
                            planned_date=first_surv_due,
                            audit_status='completed' if first_surv_cond else 'planned',
                            actual_date=first_surv_cond
                        )
                        audit._skip_cycle_creation = True
                        audit.save()
                
                # Kreiraj drugi nadzorni audit
                # VAŽNO: Onemogućavamo automatsko kreiranje resertifikacije tokom importa
                if second_surv_due:
                    existing_audit = CycleAudit.objects.filter(
                        certification_cycle=cycle,
                        audit_type='surveillance_2',
                        planned_date=second_surv_due
                    ).first()
                    
                    if existing_audit:
                        if second_surv_cond and existing_audit.actual_date != second_surv_cond:
                            existing_audit._skip_cycle_creation = True
                            existing_audit.actual_date = second_surv_cond
                            existing_audit.audit_status = 'completed'
                            existing_audit.save()
                    else:
                        audit = CycleAudit(
                            certification_cycle=cycle,
                            audit_type='surveillance_2',
                            planned_date=second_surv_due,
                            audit_status='completed' if second_surv_cond else 'planned',
                            actual_date=second_surv_cond
                        )
                        audit._skip_cycle_creation = True
                        audit.save()
                
                # Kreiraj resertifikacioni audit
                # VAŽNO: Moramo onemogućiti automatsko kreiranje novog ciklusa tokom importa
                if trinial_audit_due:
                    # Prvo proveravamo da li audit već postoji
                    existing_audit = CycleAudit.objects.filter(
                        certification_cycle=cycle,
                        audit_type='recertification',
                        planned_date=trinial_audit_due
                    ).first()
                    
                    if existing_audit:
                        # Audit već postoji, samo ažuriraj ako je potrebno
                        if trinial_audit_cond and existing_audit.actual_date != trinial_audit_cond:
                            existing_audit._skip_cycle_creation = True
                            existing_audit.actual_date = trinial_audit_cond
                            existing_audit.audit_status = 'completed'
                            existing_audit.save()
                    else:
                        # Kreiraj novi audit sa flag-om da spreči kreiranje ciklusa
                        audit = CycleAudit(
                            certification_cycle=cycle,
                            audit_type='recertification',
                            planned_date=trinial_audit_due,
                            audit_status='completed' if trinial_audit_cond else 'planned',
                            actual_date=trinial_audit_cond
                        )
                        audit._skip_cycle_creation = True
                        audit.save()
                
                created_count += 1
                
                if row_idx % 10 == 0:
                    self.stdout.write(f'  Obrađeno {row_idx} redova...')
                    
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  Greška u redu {row_idx}: {str(e)}'))
                continue
        
        self.stdout.write(self.style.SUCCESS(
            f'\n✅ Nadzorne provere: {created_count} kreirano, {skipped_count} preskočeno'
        ))
