"""
Management komanda za import duplikata kompanija iz Excel fajla.

Ova komanda importuje kompanije iz "dupli" sheeta gde više redova sa istim/sličnim 
nazivom ili PIB-om predstavljaju različite cikluse sertifikacije iste kompanije.

Logika:
1. Grupiše redove po PIB-u (ili nazivu ako PIB nije dostupan)
2. Kreira jednu kompaniju sa prvim nazivom iz grupe
3. Svaki red u grupi tretira kao zaseban ciklus sertifikacije
4. Povezuje sa auditima iz naredne-provere.xlsx koristeći company_id

Primer korišćenja:
    python manage.py import_duplicate_companies company-list.xlsx naredne-provere.xlsx
    python manage.py import_duplicate_companies company-list.xlsx naredne-provere.xlsx --dry-run
    python manage.py import_duplicate_companies company-list.xlsx naredne-provere.xlsx --limit 50
"""

from django.core.management.base import BaseCommand
from django.db import transaction
from company.models import (
    Company, CertificationCycle, CycleAudit, StandardDefinition,
    CompanyStandard, CycleStandard, CompanyIAFEACCode
)
from datetime import datetime, date
import openpyxl
from collections import defaultdict
from difflib import SequenceMatcher


class Command(BaseCommand):
    help = 'Importuje duplirane kompanije iz "dupli" sheeta i kreira multiple cikluse'

    def add_arguments(self, parser):
        parser.add_argument('company_file', type=str, help='Putanja do company-list.xlsx fajla')
        parser.add_argument('audit_file', type=str, help='Putanja do naredne-provere.xlsx fajla')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Pokreni import bez čuvanja u bazu (test mode)',
        )
        parser.add_argument(
            '--limit',
            type=int,
            help='Ograniči broj kompanija za import (za testiranje)',
        )

    def handle(self, *args, **options):
        company_file = options['company_file']
        audit_file = options['audit_file']
        dry_run = options.get('dry_run', False)
        limit = options.get('limit', None)

        if dry_run:
            self.stdout.write(self.style.WARNING('🔍 DRY RUN MODE - Podaci neće biti sačuvani u bazu'))

        # Otvori log fajl
        self.log_file = open('import_duplicates_log.txt', 'w', encoding='utf-8')
        
        try:
            with transaction.atomic():
                # Import kompanija iz "dupli" sheeta
                self.stdout.write('📂 Učitavam duplirane kompanije iz "dupli" sheeta...')
                companies_data = self.load_duplicate_companies(company_file, limit)
                
                # Grupiši po PIB-u ili sličnom nazivu
                self.stdout.write('🔗 Grupisanje kompanija po PIB-u...')
                grouped_companies = self.group_companies(companies_data)
                
                # Kreiraj kompanije i cikluse
                self.stdout.write('💾 Kreiranje kompanija i ciklusa...')
                company_id_mapping = self.create_companies_and_cycles(grouped_companies)
                
                # Import audita iz naredne-provere.xlsx
                self.stdout.write('📋 Učitavam audite iz naredne-provere.xlsx...')
                self.import_audits(audit_file, company_id_mapping)
                
                if dry_run:
                    self.stdout.write(self.style.WARNING('🔄 Rollback - Podaci nisu sačuvani'))
                    raise Exception("Dry run - rollback")
                    
                self.stdout.write(self.style.SUCCESS('✅ Import završen! Log sačuvan u import_duplicates_log.txt'))
                
        except Exception as e:
            if not dry_run:
                self.stdout.write(self.style.ERROR(f'❌ Greška: {str(e)}'))
            else:
                self.stdout.write(self.style.SUCCESS('✅ Dry run završen! Log sačuvan u import_duplicates_log.txt'))
        finally:
            self.log_file.close()

    def log(self, message):
        """Piše poruku u log fajl i terminal"""
        self.log_file.write(message + '\n')
        self.log_file.flush()

    def load_duplicate_companies(self, file_path, limit=None):
        """Učitava kompanije iz 'dupli' sheeta"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb['dupli']
        
        companies = []
        row_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if limit and row_count >= limit:
                break
                
            company_id = row[0]
            company_name = row[1]
            certificate_no = row[2]
            init_reg_date = row[3]
            standard = row[4]
            certificate_status = row[5]
            
            if not company_id or not company_name:
                continue
            
            # Parsiranje datuma
            if isinstance(init_reg_date, datetime):
                init_reg_date = init_reg_date.date()
            elif init_reg_date == '0001-01-01' or not init_reg_date:
                init_reg_date = None
            
            companies.append({
                'company_id': company_id,
                'name': str(company_name).strip(),
                'certificate_no': certificate_no,
                'init_reg_date': init_reg_date,
                'standard': standard,
                'status': certificate_status,
            })
            
            row_count += 1
            
        self.stdout.write(f'  📊 Učitano {len(companies)} redova iz "dupli" sheeta')
        self.log(f'Učitano {len(companies)} redova iz "dupli" sheeta')
        return companies

    def group_companies(self, companies_data):
        """
        Grupiše kompanije po PIB-u ili sličnom nazivu.
        Vraća dict: {group_key: [company_data1, company_data2, ...]}
        """
        groups = defaultdict(list)
        
        # Prvo grupiši po PIB-u ako postoji u nazivu
        for company in companies_data:
            # Pokušaj da ekstrahuješ PIB iz naziva ili certificate_no
            name = company['name'].upper()
            
            # Jednostavna heuristika: grupiši po osnovnom nazivu (bez PIB-a)
            # Npr. "ADAM ŠPED SYSTEM" i "ADAM-ŠPED SYSTEM DOO" → ista grupa
            base_name = self.extract_base_name(name)
            groups[base_name].append(company)
        
        self.stdout.write(f'  🔗 Grupisano u {len(groups)} grupa kompanija')
        self.log(f'Grupisano u {len(groups)} grupa kompanija')
        
        # Logiraj grupe sa više od jednog člana
        for group_key, group_companies in groups.items():
            if len(group_companies) > 1:
                self.log(f'\nGrupa "{group_key}": {len(group_companies)} ciklusa')
                for comp in group_companies:
                    self.log(f'  - ID={comp["company_id"]}, Naziv={comp["name"]}, Standard={comp["standard"]}')
        
        return groups

    def extract_base_name(self, name):
        """
        Ekstrakcija osnovnog naziva kompanije bez pravne forme i PIB-a.
        Npr: "ADAM-ŠPED SYSTEM DOO" → "ADAM ŠPED SYSTEM"
        """
        # Ukloni pravne forme
        legal_forms = ['DOO', 'D.O.O.', 'AD', 'A.D.', 'DOO BEOGRAD', 'PREDUZEĆE']
        name_upper = name.upper()
        
        for form in legal_forms:
            name_upper = name_upper.replace(form, '')
        
        # Ukloni specijalne karaktere i normalizuj razmake
        name_upper = name_upper.replace('-', ' ').replace('.', ' ')
        name_upper = ' '.join(name_upper.split())
        
        return name_upper.strip()

    def similarity(self, a, b):
        """Vraća sličnost između dva stringa (0.0 - 1.0)"""
        return SequenceMatcher(None, a.upper(), b.upper()).ratio()

    def create_companies_and_cycles(self, grouped_companies):
        """
        Kreira kompanije i cikluse iz grupisanih podataka.
        Vraća mapping: {old_company_id: (company_obj, cycle_obj)}
        """
        company_id_mapping = {}
        created_companies = 0
        created_cycles = 0
        
        for group_key, group_companies in grouped_companies.items():
            # Sortiraj po datumu (najstariji prvi)
            group_companies.sort(key=lambda x: x['init_reg_date'] if x['init_reg_date'] else date(1900, 1, 1))
            
            # Uzmi prvi red za kreiranje kompanije
            first_record = group_companies[0]
            
            # Kreiraj kompaniju (ili pronađi postojeću)
            company, created = Company.objects.get_or_create(
                name=first_record['name'],
                defaults={
                    'pib': '',
                    'certificate_status': 'active',
                    'is_active': True,
                }
            )
            
            if created:
                created_companies += 1
                self.log(f'\n✅ Kreirana kompanija: {company.name}')
            else:
                self.log(f'\n🔄 Postojeća kompanija: {company.name}')
            
            # Kreiraj ciklus za svaki red u grupi
            for record in group_companies:
                # Proveri da li već postoji ciklus sa ovim datumom
                cycle_date = record['init_reg_date'] or date.today()
                
                cycle, cycle_created = CertificationCycle.objects.get_or_create(
                    company=company,
                    planirani_datum=cycle_date,
                    defaults={
                        'status': 'active',
                        'notes': f'Importovano iz dupli sheeta, original ID: {record["company_id"]}'
                    }
                )
                
                if cycle_created:
                    created_cycles += 1
                    self.log(f'  ✅ Kreiran ciklus: datum={cycle_date}, ID={record["company_id"]}')
                
                # Dodaj standarde za ovaj ciklus
                if record['standard']:
                    self.add_standards_to_cycle(company, cycle, record['standard'], record['init_reg_date'])
                
                # Mapiranje starog ID-a na novu kompaniju i ciklus
                company_id_mapping[record['company_id']] = (company, cycle)
        
        self.stdout.write(f'  ✅ Kreirano: {created_companies} kompanija, {created_cycles} ciklusa')
        self.log(f'\n📊 Ukupno kreirano: {created_companies} kompanija, {created_cycles} ciklusa')
        
        return company_id_mapping

    def add_standards_to_cycle(self, company, cycle, standard_codes, issue_date):
        """Dodaje standarde kompaniji i ciklusu"""
        try:
            # Konvertuj u string
            standard_codes = str(standard_codes) if standard_codes else ''
            
            # Parsiranje standarda (ista logika kao u import_company_data.py)
            if standard_codes.isdigit() and len(standard_codes) > 5:
                codes = self.parse_concatenated_standards(standard_codes)
            else:
                standard_codes = standard_codes.replace('.', ',').replace(';', ',')
                if ',' in standard_codes:
                    codes = [code.strip() for code in standard_codes.split(',')]
                elif ' ' in standard_codes:
                    codes = [code.strip() for code in standard_codes.split(' ') if code.strip()]
                else:
                    codes = [standard_codes.strip()]
            
            for code in codes:
                if code:
                    self.add_standard(company, cycle, code, issue_date)
                    
        except Exception as e:
            self.log(f'    ⚠️ Greška pri dodavanju standarda: {str(e)}')

    def parse_concatenated_standards(self, standard_codes):
        """Parsira velike brojeve bez separatora (npr. 90011400118001)"""
        codes = []
        i = 0
        while i < len(standard_codes):
            # Pokušaj 5-cifrene kodove
            if i + 5 <= len(standard_codes):
                code = standard_codes[i:i+5]
                if code in ['45001', '22000', '27001', '20000', '50001', '22301']:
                    codes.append(code)
                    i += 5
                    continue
            # Proveri 4-cifrene specifične kodove
            if i + 4 <= len(standard_codes):
                code = standard_codes[i:i+4]
                if code == '3834':
                    codes.append(code)
                    i += 4
                    continue
            # Inače uzmi 4 cifre
            if i + 4 <= len(standard_codes):
                code = standard_codes[i:i+4]
                codes.append(code)
                i += 4
            else:
                codes.append(standard_codes[i:])
                break
        return codes

    def add_standard(self, company, cycle, standard_id, issue_date):
        """Dodaje jedan standard kompaniji i ciklusu"""
        try:
            # Mapiranje skraćenih kodova
            standard_mapping = {
                '9': '9001', '14': '14001', '18': '18001', '45': '45001',
                '22': '22000', '27': '27001', '20': '20000', '50': '50001',
                '3834': '3834', '22301': '22301',
            }
            
            std_code = str(standard_id).strip()
            if std_code in standard_mapping:
                std_code = standard_mapping[std_code]
            
            # Pronađi standard definiciju
            standard_def = StandardDefinition.objects.filter(
                code__icontains=std_code
            ).first()
            
            if not standard_def:
                self.log(f'    ⚠️ Standard {std_code} nije pronađen u bazi')
                return
            
            # Dodaj standard kompaniji
            CompanyStandard.objects.get_or_create(
                company=company,
                standard_definition=standard_def,
                defaults={'issue_date': issue_date}
            )
            
            # Dodaj standard ciklusu
            CycleStandard.objects.get_or_create(
                certification_cycle=cycle,
                standard_definition=standard_def
            )
            
            self.log(f'    ✅ Dodat standard: {standard_def.code}')
            
        except Exception as e:
            self.log(f'    ⚠️ Greška pri dodavanju standarda {standard_id}: {str(e)}')

    def import_audits(self, audit_file, company_id_mapping):
        """Importuje audite iz naredne-provere.xlsx za duplirane kompanije"""
        wb = openpyxl.load_workbook(audit_file)
        ws = wb.active
        
        created_audits = 0
        skipped_audits = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            company_id = row[0]
            planned_date = row[2]
            
            if not company_id or company_id not in company_id_mapping:
                skipped_audits += 1
                continue
            
            # Pronađi kompaniju i ciklus
            company, cycle = company_id_mapping[company_id]
            
            # Parsiranje datuma
            if isinstance(planned_date, datetime):
                planned_date = planned_date.date()
            elif not planned_date or planned_date == '0001-01-01':
                planned_date = date.today()
            
            # Kreiraj audit
            audit, created = CycleAudit.objects.get_or_create(
                certification_cycle=cycle,
                planned_date=planned_date,
                defaults={
                    'audit_type': 'surveillance_1',
                    'audit_status': 'completed',
                }
            )
            
            # Postavi flag da ne kreira automatski nove cikluse
            audit._skip_cycle_creation = True
            audit.save()
            
            if created:
                created_audits += 1
        
        self.stdout.write(f'  ✅ Auditi: {created_audits} kreirano, {skipped_audits} preskočeno')
        self.log(f'\n📊 Auditi: {created_audits} kreirano, {skipped_audits} preskočeno')
