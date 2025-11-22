"""
Management komanda za import audita iz naredne-provere.xlsx samo za duplirane kompanije.

Ova komanda:
1. Učitava company_id iz "dupli" sheeta
2. Filtrira naredne-provere.xlsx samo za te company_id
3. Uparuje company_id sa kompanijama koje su već importovane
4. Kreira audite za odgovarajuće cikluse

Primer korišćenja:
    python manage.py import_duplicate_audits company-list.xlsx naredne-provere.xlsx
    python manage.py import_duplicate_audits company-list.xlsx naredne-provere.xlsx --dry-run
"""

from django.core.management.base import BaseCommand
from django.db import transaction
from company.models import Company, CertificationCycle, CycleAudit
from datetime import datetime, date
import openpyxl


class Command(BaseCommand):
    help = 'Importuje audite iz naredne-provere.xlsx samo za duplirane kompanije'

    def add_arguments(self, parser):
        parser.add_argument('company_file', type=str, help='Putanja do company-list.xlsx fajla')
        parser.add_argument('audit_file', type=str, help='Putanja do naredne-provere.xlsx fajla')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Pokreni import bez čuvanja u bazu (test mode)',
        )

    def handle(self, *args, **options):
        company_file = options['company_file']
        audit_file = options['audit_file']
        dry_run = options.get('dry_run', False)

        if dry_run:
            self.stdout.write(self.style.WARNING('🔍 DRY RUN MODE - Podaci neće biti sačuvani u bazu'))

        # Otvori log fajl
        self.log_file = open('import_duplicate_audits_log.txt', 'w', encoding='utf-8')
        
        try:
            with transaction.atomic():
                # 1. Učitaj company_id iz "dupli" sheeta
                self.stdout.write('📂 Učitavam company_id iz "dupli" sheeta...')
                duplicate_company_ids = self.load_duplicate_company_ids(company_file)
                
                # 2. Kreiraj mapiranje company_id -> Company objekat
                self.stdout.write('🔗 Kreiram mapiranje company_id -> Company...')
                company_mapping = self.create_company_mapping(duplicate_company_ids)
                
                # 3. Importuj audite samo za te company_id
                self.stdout.write('📋 Importujem audite iz naredne-provere.xlsx...')
                self.import_audits(audit_file, company_mapping)
                
                if dry_run:
                    self.stdout.write(self.style.WARNING('🔄 Rollback - Podaci nisu sačuvani'))
                    raise Exception("Dry run - rollback")
                    
                self.stdout.write(self.style.SUCCESS('✅ Import završen! Log sačuvan u import_duplicate_audits_log.txt'))
                
        except Exception as e:
            if not dry_run:
                self.stdout.write(self.style.ERROR(f'❌ Greška: {str(e)}'))
            else:
                self.stdout.write(self.style.SUCCESS('✅ Dry run završen! Log sačuvan u import_duplicate_audits_log.txt'))
        finally:
            self.log_file.close()

    def log(self, message):
        """Piše poruku u log fajl i terminal"""
        self.log_file.write(message + '\n')
        self.log_file.flush()

    def load_duplicate_company_ids(self, file_path):
        """Učitava sve company_id iz 'dupli' sheeta"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb['dupli']
        
        company_ids = set()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            company_id = row[0]
            if company_id:
                company_ids.add(company_id)
        
        self.stdout.write(f'  📊 Pronađeno {len(company_ids)} jedinstvenih company_id u "dupli" sheetu')
        self.log(f'Pronađeno {len(company_ids)} jedinstvenih company_id: {sorted(company_ids)[:20]}...')
        
        return company_ids

    def create_company_mapping(self, duplicate_company_ids):
        """
        Kreira mapiranje: company_id -> (Company, CertificationCycle)
        
        Logika:
        - Za svaki company_id iz "dupli" sheeta, pronađi odgovarajući ciklus
        - Svaki company_id ima svoj zaseban ciklus
        """
        company_mapping = {}
        
        self.log(f'\nMapiranje company_id -> (Company, CertificationCycle):')
        
        for company_id in duplicate_company_ids:
            # Pronađi ciklus koji ima ovaj company_id u notes
            cycle = CertificationCycle.objects.filter(
                notes__icontains=f'original ID: {company_id}'
            ).select_related('company').first()
            
            if cycle:
                company = cycle.company
                company_mapping[company_id] = (company, cycle)
                self.log(f'  ✅ ID {company_id} -> {company.name} (ciklus: {cycle.planirani_datum})')
            else:
                self.log(f'  ⚠️ ID {company_id} -> Ciklus nije pronađen')
        
        self.stdout.write(f'  🔗 Mapirano {len(company_mapping)} company_id na cikluse')
        self.log(f'\nUkupno mapirano: {len(company_mapping)} company_id')
        
        return company_mapping

    def import_audits(self, audit_file, company_mapping):
        """Importuje audite samo za kompanije iz company_mapping"""
        wb = openpyxl.load_workbook(audit_file)
        ws = wb.active
        
        created_audits = 0
        skipped_audits = 0
        not_found = 0
        
        self.log(f'\nImport audita:')
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            company_id = row[1]  # Kolona 'company_id' je na indeksu 1
            
            # Preskoči ako company_id nije u našem mappingu
            if not company_id or company_id not in company_mapping:
                if company_id:
                    not_found += 1
                    if not_found <= 10:  # Logiraj samo prvih 10
                        self.log(f'  Red {row_num}: company_id={company_id} nije u dupli sheetu - preskačem')
                skipped_audits += 1
                continue
            
            # Pronađi kompaniju i ciklus
            company, cycle = company_mapping[company_id]
            
            # Izvuci sve datume audita iz reda
            # Kolone: [0]=naredne_provere_id, [1]=company_id, [2]=first_surv_due, [3]=first_surv_cond,
            #         [4]=second_surv_due, [5]=second_surv_cond, [6]=trinial_audit_due, [7]=trinial_audit_cond, [8]=status_id
            audits_to_create = [
                {'planned': row[2], 'actual': row[3], 'type': 'surveillance_1', 'name': '1st Surveillance'},
                {'planned': row[4], 'actual': row[5], 'type': 'surveillance_2', 'name': '2nd Surveillance'},
                {'planned': row[6], 'actual': row[7], 'type': 'recertification', 'name': 'Trianial/Recertification'},
            ]
            
            for audit_data in audits_to_create:
                planned_date = audit_data['planned']
                actual_date = audit_data['actual']
                audit_type = audit_data['type']
                audit_name = audit_data['name']
                
                # Preskoči ako nema planirani datum
                if not planned_date:
                    continue
                
                # Parsiranje planiranog datuma
                if isinstance(planned_date, datetime):
                    planned_date = planned_date.date()
                elif str(planned_date) == '0001-01-01':
                    continue
                
                # Parsiranje stvarnog datuma
                if actual_date and isinstance(actual_date, datetime):
                    actual_date = actual_date.date()
                elif not actual_date or str(actual_date) == '0001-01-01':
                    actual_date = None
                
                # Odredi status audita
                audit_status = 'completed' if actual_date else 'planned'
                
                # Kreiraj audit
                try:
                    audit, created = CycleAudit.objects.get_or_create(
                        certification_cycle=cycle,
                        planned_date=planned_date,
                        audit_type=audit_type,
                        defaults={
                            'audit_status': audit_status,
                            'actual_date': actual_date,
                        }
                    )
                    
                    # Postavi flag da ne kreira automatski nove cikluse
                    if created:
                        audit._skip_cycle_creation = True
                        audit.save()
                        created_audits += 1
                        status_str = f'completed ({actual_date})' if actual_date else 'planned'
                        self.log(f'  ✅ Red {row_num}: Kreiran {audit_name} za {company.name}, planirano: {planned_date}, status: {status_str}')
                    else:
                        skipped_audits += 1
                        
                except Exception as e:
                    self.log(f'  ❌ Red {row_num}: Greška pri kreiranju {audit_name}: {str(e)}')
                    skipped_audits += 1
        
        self.stdout.write(f'  ✅ Auditi: {created_audits} kreirano, {skipped_audits} preskočeno')
        self.log(f'\n📊 Ukupno: {created_audits} audita kreirano, {skipped_audits} preskočeno, {not_found} nije pronađeno u dupli sheetu')
