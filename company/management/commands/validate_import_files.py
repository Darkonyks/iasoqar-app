"""
Django management command za validaciju Excel fajlova pre importa
"""
from django.core.management.base import BaseCommand
from company.models import StandardDefinition, IAFEACCode
import openpyxl
import os
from datetime import datetime


class Command(BaseCommand):
    help = 'Validira Excel fajlove pre importa podataka'

    def add_arguments(self, parser):
        parser.add_argument(
            'company_file',
            type=str,
            help='Putanja do Excel fajla sa kompanijama (company-list.xlsx)'
        )
        parser.add_argument(
            'provere_file',
            type=str,
            help='Putanja do Excel fajla sa nadzornim proverama (naredne-provere.xlsx)'
        )

    def handle(self, *args, **options):
        company_file = options['company_file']
        provere_file = options['provere_file']

        self.stdout.write(self.style.SUCCESS('🔍 Validacija Excel fajlova...\n'))

        # Validacija company fajla
        company_errors = self.validate_company_file(company_file)
        
        # Validacija provere fajla
        provere_errors = self.validate_provere_file(provere_file, company_file)
        
        # Prikaz rezultata
        total_errors = len(company_errors) + len(provere_errors)
        
        if total_errors == 0:
            self.stdout.write(self.style.SUCCESS('\n✅ Validacija uspešna! Fajlovi su spremni za import.'))
        else:
            self.stdout.write(self.style.ERROR(f'\n❌ Pronađeno {total_errors} grešaka!'))
            self.stdout.write(self.style.WARNING('\nMolimo ispravite greške pre importa.'))

    def validate_company_file(self, file_path):
        """Validira company-list.xlsx fajl"""
        self.stdout.write(self.style.SUCCESS(f'📊 Validacija: {file_path}'))
        
        errors = []
        
        if not os.path.exists(file_path):
            errors.append(f'Fajl ne postoji: {file_path}')
            self.stdout.write(self.style.ERROR(f'  ❌ Fajl ne postoji'))
            return errors
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Proveri header
            headers = [cell.value for cell in ws[1]]
            expected_headers = [
                'company_id', 'company_name', 'certificate_number', 'IAF_kod_id',
                'standard_id', 'certificate_start', 'suspension_until_date',
                'audit_days', 'visits_per_year', 'audit_days_each', 'contracted_date'
            ]
            
            self.stdout.write(f'  Kolone: {headers}')
            
            # Validacija podataka
            company_ids = set()
            missing_names = []
            invalid_standards = []
            invalid_iaf_codes = []
            invalid_dates = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                company_id = row[0]
                company_name = row[1]
                standard_id = row[4] if len(row) > 4 else None
                iaf_kod_id = row[3] if len(row) > 3 else None
                certificate_start = row[5] if len(row) > 5 else None
                
                # Proveri duplikate company_id
                if company_id in company_ids:
                    errors.append(f'Red {row_idx}: Duplikat company_id: {company_id}')
                company_ids.add(company_id)
                
                # Proveri da li postoji naziv
                if not company_name:
                    missing_names.append(row_idx)
                
                # Proveri da li standard postoji u bazi
                if standard_id:
                    if not self.standard_exists(standard_id):
                        invalid_standards.append((row_idx, standard_id))
                
                # Proveri da li IAF kod postoji u bazi
                if iaf_kod_id:
                    if not self.iaf_code_exists(iaf_kod_id):
                        invalid_iaf_codes.append((row_idx, iaf_kod_id))
                
                # Proveri format datuma
                if certificate_start and not self.is_valid_date(certificate_start):
                    invalid_dates.append((row_idx, 'certificate_start', certificate_start))
            
            # Prikaz rezultata
            self.stdout.write(f'  ✅ Ukupno redova: {len(company_ids)}')
            
            if missing_names:
                self.stdout.write(self.style.WARNING(
                    f'  ⚠️  Redovi bez naziva kompanije: {missing_names}'
                ))
                errors.extend([f'Red {r}: Nedostaje naziv kompanije' for r in missing_names])
            
            if invalid_standards:
                self.stdout.write(self.style.WARNING(
                    f'  ⚠️  Nepostojeći standardi: {len(invalid_standards)}'
                ))
                for row, std_id in invalid_standards[:5]:  # Prikaži prvih 5
                    self.stdout.write(f'     Red {row}: Standard {std_id} ne postoji u bazi')
                    errors.append(f'Red {row}: Standard {std_id} ne postoji')
            
            if invalid_iaf_codes:
                self.stdout.write(self.style.WARNING(
                    f'  ⚠️  Nepostojeći IAF kodovi: {len(invalid_iaf_codes)}'
                ))
                for row, iaf_id in invalid_iaf_codes[:5]:  # Prikaži prvih 5
                    self.stdout.write(f'     Red {row}: IAF kod {iaf_id} ne postoji u bazi')
                    errors.append(f'Red {row}: IAF kod {iaf_id} ne postoji')
            
            if invalid_dates:
                self.stdout.write(self.style.WARNING(
                    f'  ⚠️  Nevalidni datumi: {len(invalid_dates)}'
                ))
                for row, field, value in invalid_dates[:5]:
                    self.stdout.write(f'     Red {row}: {field} = {value}')
                    errors.append(f'Red {row}: Nevalidan datum {field}: {value}')
            
            if not errors:
                self.stdout.write(self.style.SUCCESS('  ✅ Sve validacije prošle!'))
            
        except Exception as e:
            errors.append(f'Greška pri čitanju fajla: {str(e)}')
            self.stdout.write(self.style.ERROR(f'  ❌ Greška: {str(e)}'))
        
        return errors

    def validate_provere_file(self, file_path, company_file_path):
        """Validira naredne-provere.xlsx fajl"""
        self.stdout.write(self.style.SUCCESS(f'\n📊 Validacija: {file_path}'))
        
        errors = []
        
        if not os.path.exists(file_path):
            errors.append(f'Fajl ne postoji: {file_path}')
            self.stdout.write(self.style.ERROR(f'  ❌ Fajl ne postoji'))
            return errors
        
        try:
            # Učitaj company IDs iz company fajla
            company_ids = self.get_company_ids(company_file_path)
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Proveri header
            headers = [cell.value for cell in ws[1]]
            self.stdout.write(f'  Kolone: {headers}')
            
            # Validacija podataka
            missing_companies = []
            invalid_dates = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                company_id = row[1]
                first_surv_due = row[2] if len(row) > 2 else None
                first_surv_cond = row[3] if len(row) > 3 else None
                second_surv_due = row[4] if len(row) > 4 else None
                second_surv_cond = row[5] if len(row) > 5 else None
                trinial_audit_due = row[6] if len(row) > 6 else None
                trinial_audit_cond = row[7] if len(row) > 7 else None
                
                # Proveri da li company_id postoji
                if company_id not in company_ids:
                    missing_companies.append((row_idx, company_id))
                
                # Proveri datume
                date_fields = [
                    ('first_surv_due', first_surv_due),
                    ('first_surv_cond', first_surv_cond),
                    ('second_surv_due', second_surv_due),
                    ('second_surv_cond', second_surv_cond),
                    ('trinial_audit_due', trinial_audit_due),
                    ('trinial_audit_cond', trinial_audit_cond),
                ]
                
                for field_name, field_value in date_fields:
                    if field_value and not self.is_valid_date(field_value):
                        invalid_dates.append((row_idx, field_name, field_value))
            
            # Prikaz rezultata
            total_rows = ws.max_row - 1
            self.stdout.write(f'  ✅ Ukupno redova: {total_rows}')
            
            if missing_companies:
                self.stdout.write(self.style.WARNING(
                    f'  ⚠️  Nepostojeće kompanije: {len(missing_companies)}'
                ))
                for row, comp_id in missing_companies[:5]:
                    self.stdout.write(f'     Red {row}: Kompanija ID {comp_id} ne postoji u company-list.xlsx')
                    errors.append(f'Red {row}: Kompanija ID {comp_id} ne postoji')
            
            if invalid_dates:
                self.stdout.write(self.style.WARNING(
                    f'  ⚠️  Nevalidni datumi: {len(invalid_dates)}'
                ))
                for row, field, value in invalid_dates[:5]:
                    self.stdout.write(f'     Red {row}: {field} = {value}')
                    errors.append(f'Red {row}: Nevalidan datum {field}: {value}')
            
            if not errors:
                self.stdout.write(self.style.SUCCESS('  ✅ Sve validacije prošle!'))
            
        except Exception as e:
            errors.append(f'Greška pri čitanju fajla: {str(e)}')
            self.stdout.write(self.style.ERROR(f'  ❌ Greška: {str(e)}'))
        
        return errors

    def get_company_ids(self, file_path):
        """Vraća set svih company_id iz company fajla"""
        company_ids = set()
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                company_id = row[0]
                if company_id:
                    company_ids.add(company_id)
        except:
            pass
        
        return company_ids

    def standard_exists(self, standard_id):
        """Proveri da li standard postoji u bazi"""
        try:
            if isinstance(standard_id, int) or (isinstance(standard_id, str) and standard_id.isdigit()):
                return StandardDefinition.objects.filter(id=int(standard_id)).exists()
            else:
                return StandardDefinition.objects.filter(code__icontains=str(standard_id)).exists()
        except:
            return False

    def iaf_code_exists(self, iaf_kod_id):
        """Proveri da li IAF kod postoji u bazi"""
        try:
            if isinstance(iaf_kod_id, int) or (isinstance(iaf_kod_id, str) and iaf_kod_id.isdigit()):
                return IAFEACCode.objects.filter(id=int(iaf_kod_id)).exists()
            else:
                return IAFEACCode.objects.filter(code__icontains=str(iaf_kod_id)).exists()
        except:
            return False

    def is_valid_date(self, date_value):
        """Proveri da li je datum validan"""
        if not date_value:
            return True
        
        if isinstance(date_value, datetime):
            return True
        
        if isinstance(date_value, str):
            for fmt in ['%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                try:
                    datetime.strptime(date_value, fmt)
                    return True
                except ValueError:
                    continue
        
        return False
