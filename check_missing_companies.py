"""
Skripta za proveru nedostajućih kompanija između dva Excel fajla
"""
import openpyxl
from collections import defaultdict

def main():
    # Učitaj company IDs iz company-list.xlsx
    print("📊 Učitavanje company-list.xlsx...")
    wb_companies = openpyxl.load_workbook('company-list.xlsx')
    ws_companies = wb_companies.active
    
    company_ids = set()
    company_names = {}
    
    for row in ws_companies.iter_rows(min_row=2, values_only=True):
        if row[0]:
            company_ids.add(row[0])
            company_names[row[0]] = row[1] if len(row) > 1 else 'N/A'
    
    print(f"✅ Učitano {len(company_ids)} kompanija\n")
    
    # Učitaj company IDs iz naredne-provere.xlsx
    print("📊 Učitavanje naredne-provere.xlsx...")
    wb_provere = openpyxl.load_workbook('naredne-provere.xlsx')
    ws_provere = wb_provere.active
    
    missing_companies = defaultdict(list)
    total_provere = 0
    
    for row_idx, row in enumerate(ws_provere.iter_rows(min_row=2, values_only=True), start=2):
        total_provere += 1
        company_id = row[1] if len(row) > 1 else None
        
        if company_id and company_id not in company_ids:
            # Čuvaj informacije o proveri
            provera_info = {
                'row': row_idx,
                'naredne_provere_id': row[0],
                'first_surv_due': row[2] if len(row) > 2 else None,
                'first_surv_cond': row[3] if len(row) > 3 else None,
                'second_surv_due': row[4] if len(row) > 4 else None,
                'second_surv_cond': row[5] if len(row) > 5 else None,
            }
            missing_companies[company_id].append(provera_info)
    
    print(f"✅ Učitano {total_provere} provera\n")
    
    # Prikaži rezultate
    print("="*80)
    print("📊 IZVEŠTAJ O NEDOSTAJUĆIM KOMPANIJAMA")
    print("="*80)
    print(f"\nUkupno kompanija u company-list.xlsx: {len(company_ids)}")
    print(f"Ukupno provera u naredne-provere.xlsx: {total_provere}")
    print(f"Ukupno nedostajućih kompanija: {len(missing_companies)}")
    print(f"Ukupno provera za nedostajuće kompanije: {sum(len(v) for v in missing_companies.values())}")
    
    print("\n" + "="*80)
    print("❌ LISTA NEDOSTAJUĆIH KOMPANIJA")
    print("="*80)
    
    # Sortiraj po company_id
    for comp_id in sorted(missing_companies.keys()):
        provere = missing_companies[comp_id]
        print(f"\n🔴 Company ID: {comp_id}")
        print(f"   Broj provera: {len(provere)}")
        print(f"   Redovi u naredne-provere.xlsx: {[p['row'] for p in provere]}")
        
        # Prikaži prvu proveru kao primer
        if provere:
            first = provere[0]
            print(f"   Primer provere:")
            print(f"     - Prva nadzorna (due): {first['first_surv_due']}")
            print(f"     - Prva nadzorna (cond): {first['first_surv_cond']}")
            print(f"     - Druga nadzorna (due): {first['second_surv_due']}")
    
    # Sačuvaj u tekstualni fajl
    with open('missing_companies_report.txt', 'w', encoding='utf-8') as f:
        f.write("IZVEŠTAJ O NEDOSTAJUĆIM KOMPANIJAMA\n")
        f.write("="*80 + "\n\n")
        f.write(f"Ukupno kompanija u company-list.xlsx: {len(company_ids)}\n")
        f.write(f"Ukupno provera u naredne-provere.xlsx: {total_provere}\n")
        f.write(f"Ukupno nedostajućih kompanija: {len(missing_companies)}\n")
        f.write(f"Ukupno provera za nedostajuće kompanije: {sum(len(v) for v in missing_companies.values())}\n\n")
        
        f.write("LISTA NEDOSTAJUĆIH COMPANY ID-eva:\n")
        f.write("-"*80 + "\n")
        
        for comp_id in sorted(missing_companies.keys()):
            provere = missing_companies[comp_id]
            f.write(f"\nCompany ID: {comp_id}\n")
            f.write(f"  Broj provera: {len(provere)}\n")
            f.write(f"  Redovi: {[p['row'] for p in provere]}\n")
    
    print("\n" + "="*80)
    print(f"✅ Izveštaj sačuvan u: missing_companies_report.txt")
    print("="*80)

if __name__ == '__main__':
    main()
